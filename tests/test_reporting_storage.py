from pathlib import Path

import pytest

from m365_assessor.core.cache import write_evidence_cache
from m365_assessor.models.assessment import (
    AssessmentDocument,
    AssessmentMetadata,
    AssessmentSummary,
    AuthenticationSummary,
    CheckResult,
    CollectorExecution,
    CoverageArea,
    TenantInfo,
)
from m365_assessor.models.enums import (
    AssessmentStatus,
    CollectorStatus,
    CoverageStatus,
    Severity,
)
from m365_assessor.reporting.generator import generate_reports
from m365_assessor.reporting.json_report import write_json_report
from m365_assessor.storage.repository import AssessmentRepository


def _document() -> AssessmentDocument:
    return AssessmentDocument(
        assessment=AssessmentMetadata(),
        tenant=TenantInfo(tenant_id="tenant"),
        authentication=AuthenticationSummary(
            method="interactive",
            tenant_id="tenant",
            client_id="client",
            permission_source="token_response",
        ),
        coverage=[],
        collectors={},
        checks=[],
        findings=[],
        summary=AssessmentSummary(),
        benchmark_results={},
    )


def test_json_round_trip_and_no_token(tmp_path: Path) -> None:
    destination = write_json_report(_document(), tmp_path)
    content = destination.read_text(encoding="utf-8")
    assert "access_token" not in content
    assert AssessmentDocument.model_validate_json(content).tenant.tenant_id == "tenant"


def test_json_filename_rejects_traversal(tmp_path: Path) -> None:
    with pytest.raises(ValueError):
        write_json_report(_document(), tmp_path, "../escape.json")


def test_sqlalchemy_repository_round_trip(tmp_path: Path) -> None:
    repository = AssessmentRepository(f"sqlite:///{tmp_path / 'test.db'}")
    repository.initialize()
    document = _document()
    repository.save(document)
    loaded = repository.get(str(document.assessment.id))
    assert loaded is not None
    assert loaded.tenant.tenant_id == "tenant"


def test_evidence_cache_uses_area_and_collector_boundaries(tmp_path: Path) -> None:
    document = _document()
    document.collectors = {
        "entra001": CollectorExecution(
            collector_id="entra001",
            name="Tenant",
            area="entra",
            status=CollectorStatus.SUCCESS,
            data={"organization": {"id": "tenant"}},
        )
    }
    scan = write_evidence_cache(document, tmp_path)
    assert scan.name == "scan.json"
    assert (tmp_path / "evidence" / "entra" / "entra001.json").exists()


def test_all_report_formats_are_generated_and_readable(tmp_path: Path) -> None:
    document = _document()
    failed = CheckResult(
        check_id="TEST-FAIL",
        title="Escaped <finding>",
        severity=Severity.HIGH,
        status=AssessmentStatus.FAIL,
        service="Entra ID",
        category="Identity",
        observation="Condition failed.",
        remediation="Apply the approved setting.",
        evidence={"unsafe": "<script>alert(1)</script>"},
    )
    passed = failed.model_copy(
        update={"check_id": "TEST-PASS", "title": "Good control", "status": AssessmentStatus.PASS}
    )
    unavailable = failed.model_copy(
        update={
            "check_id": "TEST-NA",
            "title": "Unavailable",
            "status": AssessmentStatus.NOT_ASSESSED,
        }
    )
    document.checks = [failed, passed, unavailable]
    document.findings = [failed]
    document.summary = AssessmentSummary.from_results(document.checks)
    document.coverage = [
        CoverageArea(
            area="entra",
            name="Entra ID",
            status=CoverageStatus.PARTIAL,
            coverage_percent=66.67,
            reasons=["One check was not assessed."],
        )
    ]
    paths = generate_reports(document, tmp_path)
    assert set(paths) == {"json", "html", "csv", "xlsx"}
    assert all(path.exists() and path.stat().st_size > 0 for path in paths.values())
    html = paths["html"].read_text(encoding="utf-8")
    assert "Escaped &lt;finding&gt;" in html
    assert "<script>alert(1)</script>" not in html

    from openpyxl import load_workbook

    workbook = load_workbook(paths["xlsx"], data_only=False)
    assert workbook.sheetnames == [
        "Executive Summary",
        "Finding Register",
        "Good Controls",
        "Not Assessed",
        "Permission Coverage",
        "Framework Mapping",
        "Raw Evidence",
        "Methodology",
    ]
    assert workbook["Finding Register"]["A4"].value == "TEST-FAIL"
    for sheet_name in workbook.sheetnames[1:]:
        sheet = workbook[sheet_name]
        assert sheet.auto_filter.ref is None
        assert len(sheet.tables) == 1
        table = next(iter(sheet.tables.values()))
        assert table.autoFilter is not None
        assert table.autoFilter.ref == table.ref


def test_reports_support_a_client_timestamp_filename_stem(tmp_path: Path) -> None:
    stem = "Contoso-m365-assessment-20260819-164510"
    paths = generate_reports(_document(), tmp_path, filename_stem=stem)
    assert paths["json"].name == f"{stem}.json"
    assert paths["html"].name == f"{stem}.html"
    assert paths["xlsx"].name == f"{stem}.xlsx"
    assert paths["csv"].name == f"{stem}-findings.csv"
