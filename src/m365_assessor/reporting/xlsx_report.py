from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from m365_assessor.models.assessment import AssessmentDocument, CheckResult
from m365_assessor.reporting.io import atomic_report_path, validate_report_filename

NAVY = "17324D"
BLUE = "0B6E99"
WHITE = "FFFFFF"
LIGHT_BLUE = "DCEEF7"
GREEN = "D9EAD3"
RED = "F4CCCC"
AMBER = "FCE5CD"
GRAY = "E7E6E6"
SEVERITY_COLORS = {
    "Critical": "8B0000",
    "High": "C00000",
    "Medium": "F4B183",
    "Low": "FFD966",
    "Informational": "9DC3E6",
}


def _title(sheet: Worksheet, text: str, end_column: int) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = sheet.cell(1, 1, text)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(color=WHITE, bold=True, size=16)
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28


def _table(
    sheet: Worksheet, headers: list[str], rows: list[list[Any]], name: str, start_row: int = 3
) -> None:
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(start_row, column, header)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    safe_rows = rows or [["" for _ in headers]]
    for row_index, data_row in enumerate(safe_rows, start_row + 1):
        for column, value in enumerate(data_row, 1):
            cell = sheet.cell(row_index, column, value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    end_row = start_row + len(safe_rows)
    end_column = len(headers)
    reference = f"A{start_row}:{sheet.cell(end_row, end_column).coordinate}"
    table = Table(displayName=name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True
    )
    sheet.add_table(table)
    sheet.freeze_panes = f"A{start_row + 1}"
    # The table owns its AutoFilter. A second worksheet-level AutoFilter over
    # the same range is tolerated by openpyxl but repaired as corrupt by Excel.
    for column, header in enumerate(headers, 1):
        width = min(max(len(header) + 2, 13), 45)
        for row_index in range(start_row + 1, end_row + 1):
            value = sheet.cell(row_index, column).value
            if value is not None:
                width = min(max(width, len(str(value)) + 2), 45)
        sheet.column_dimensions[get_column_letter(column)].width = width


def _mapping_text(check: CheckResult) -> str:
    return "; ".join(f"{item.framework} {item.version} {item.control}" for item in check.benchmark)


def _check_rows(checks: list[CheckResult]) -> list[list[Any]]:
    return [
        [
            item.check_id,
            item.title,
            item.severity.value,
            item.status.value,
            item.service,
            item.category,
            _mapping_text(item),
            item.observation,
            item.business_impact,
            item.recommendation,
            item.remediation,
            ", ".join(item.affected_resources),
            item.references[0] if item.references else "",
            item.timestamp.replace(tzinfo=None),
        ]
        for item in checks
    ]


def _style_check_sheet(sheet: Worksheet, row_count: int) -> None:
    if row_count <= 0:
        return
    end = row_count + 3
    for severity, color in SEVERITY_COLORS.items():
        sheet.conditional_formatting.add(
            f"C4:C{end}",
            FormulaRule(formula=[f'$C4="{severity}"'], fill=PatternFill("solid", fgColor=color)),
        )
    for row in range(4, end + 1):
        reference = sheet.cell(row, 13)
        if reference.value:
            reference.hyperlink = str(reference.value)
            reference.style = "Hyperlink"
        sheet.cell(row, 14).number_format = "yyyy-mm-dd hh:mm"


def write_xlsx_report(
    document: AssessmentDocument,
    output_directory: Path,
    filename: str = "assessment.xlsx",
) -> Path:
    validate_report_filename(filename, ".xlsx")
    output_directory.mkdir(parents=True, exist_ok=True)
    destination = output_directory.resolve() / filename
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Executive Summary"
    _title(summary_sheet, "Microsoft 365 Security Assessment", 8)
    framework = ", ".join(f"{item.name} {item.version}" for item in document.assessment.frameworks)
    assessment_date = document.assessment.completed_at or document.assessment.started_at
    assessment_date = assessment_date.replace(tzinfo=None)
    summary_rows: list[tuple[str, Any]] = [
        ("Client", document.assessment.client_name or "Not specified"),
        ("Tenant", document.tenant.display_name or document.tenant.tenant_id),
        ("Primary domain", document.tenant.primary_domain or "Not discovered"),
        ("Assessment date", assessment_date),
        ("Frameworks", framework),
        ("Scope", ", ".join(document.assessment.scope)),
        ("Total checks", document.summary.total_checks),
        ("Passed", document.summary.pass_count),
        ("Failed", document.summary.fail_count),
        ("Warnings", document.summary.warning_count),
        ("Not assessed", document.summary.not_assessed_count),
        ("Errors", document.summary.error_count),
        ("Pass percentage", document.summary.pass_percentage / 100),
        ("Coverage percentage", document.summary.coverage_percentage / 100),
    ]
    for row, (label, value) in enumerate(summary_rows, 3):
        summary_sheet.cell(row, 1, label).font = Font(bold=True, color=NAVY)
        summary_sheet.cell(row, 2, value)
    for row in (15, 16):
        summary_sheet.cell(row, 2).number_format = "0.0%"
    summary_sheet.column_dimensions["A"].width = 24
    summary_sheet.column_dimensions["B"].width = 65
    summary_sheet.freeze_panes = "A3"
    chart_data: list[list[Any]] = [
        ["Severity", "Findings"],
        *[[key, value] for key, value in document.summary.severity_distribution.items()],
    ]
    for chart_row in chart_data:
        summary_sheet.append(chart_row)
    chart = BarChart()
    chart.title = "Finding severity"
    chart.y_axis.title = "Findings"
    chart.add_data(
        Reference(summary_sheet, min_col=2, min_row=17, max_row=22), titles_from_data=True
    )
    chart.set_categories(Reference(summary_sheet, min_col=1, min_row=18, max_row=22))
    chart.height = 7
    chart.width = 13
    summary_sheet.add_chart(chart, "D3")

    headers = [
        "Finding ID",
        "Title",
        "Severity",
        "Status",
        "Service",
        "Category",
        "Framework mapping",
        "Observation",
        "Business impact",
        "Recommendation",
        "Remediation",
        "Affected resources",
        "Reference",
        "Timestamp",
    ]
    finding_sheet = workbook.create_sheet("Finding Register")
    _title(finding_sheet, "Finding Register", len(headers))
    _table(finding_sheet, headers, _check_rows(document.findings), "FindingRegister")
    _style_check_sheet(finding_sheet, len(document.findings))

    good = [item for item in document.checks if item.status.value == "PASS"]
    good_sheet = workbook.create_sheet("Good Controls")
    _title(good_sheet, "Good Controls", len(headers))
    _table(good_sheet, headers, _check_rows(good), "GoodControls")
    _style_check_sheet(good_sheet, len(good))

    unavailable = [
        item for item in document.checks if item.status.value in {"NOT_ASSESSED", "ERROR"}
    ]
    na_sheet = workbook.create_sheet("Not Assessed")
    _title(na_sheet, "Not Assessed and Errors", len(headers))
    _table(na_sheet, headers, _check_rows(unavailable), "NotAssessed")
    _style_check_sheet(na_sheet, len(unavailable))

    coverage_sheet = workbook.create_sheet("Permission Coverage")
    coverage_headers = [
        "Area",
        "Status",
        "Coverage",
        "Missing permissions",
        "Affected collectors",
        "Affected checks",
        "Explanation",
    ]
    coverage_rows = [
        [
            area.name,
            area.status.value,
            area.coverage_percent / 100,
            ", ".join(area.missing_permissions),
            ", ".join(area.affected_collectors),
            ", ".join(area.affected_checks),
            "; ".join(area.reasons),
        ]
        for area in document.coverage
    ]
    _title(coverage_sheet, "Coverage and Permission Report", len(coverage_headers))
    _table(coverage_sheet, coverage_headers, coverage_rows, "PermissionCoverage")
    for row in range(4, 4 + len(coverage_rows)):
        coverage_sheet.cell(row, 3).number_format = "0.0%"

    mapping_sheet = workbook.create_sheet("Framework Mapping")
    mapping_headers = [
        "Framework",
        "Benchmark",
        "Version",
        "Control",
        "Check ID",
        "Title",
        "Status",
        "Severity",
    ]
    mapping_rows = [
        [
            mapping.framework,
            mapping.benchmark,
            mapping.version,
            mapping.control,
            check.check_id,
            check.title,
            check.status.value,
            check.severity.value,
        ]
        for check in document.checks
        for mapping in check.benchmark
    ]
    _title(mapping_sheet, "Framework Mapping", len(mapping_headers))
    _table(mapping_sheet, mapping_headers, mapping_rows, "FrameworkMapping")

    evidence_sheet = workbook.create_sheet("Raw Evidence")
    evidence_headers = ["Collector ID", "Name", "Area", "Status", "Objects", "Evidence JSON"]
    evidence_rows = [
        [
            item.collector_id,
            item.name,
            item.area,
            item.status.value,
            item.objects_collected,
            json.dumps(item.data, ensure_ascii=False, default=str)[:32700],
        ]
        for item in document.collectors.values()
    ]
    _title(evidence_sheet, "Raw Evidence", len(evidence_headers))
    _table(evidence_sheet, evidence_headers, evidence_rows, "RawEvidence")

    methodology_sheet = workbook.create_sheet("Methodology")
    _title(methodology_sheet, "Methodology and Limitations", 4)
    methodology_rows = [
        ["Principle", "Read-only collection; the assessment never changes tenant configuration."],
        [
            "Collection",
            "Microsoft Graph and Microsoft-supported administrative PowerShell modules.",
        ],
        ["Evaluation", "Deterministic versioned YAML rules evaluated against normalized evidence."],
        ["Permission gaps", "Unavailable evidence is NOT_ASSESSED and is never treated as FAIL."],
        ["Scoring", "Pass percentage excludes NOT_ASSESSED, NOT_APPLICABLE, and ERROR controls."],
        [
            "Sensitive data",
            "Tokens, passwords, secrets, and private keys are excluded from reports.",
        ],
    ]
    _table(methodology_sheet, ["Topic", "Method"], methodology_rows, "Methodology")

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    with atomic_report_path(destination) as temporary:
        workbook.save(temporary)
    return destination
